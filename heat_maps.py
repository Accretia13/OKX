import pandas as pd
import numpy as np
import os
import glob
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def col_letter(n):
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

OTHER_TF = r"C:\Users\777\PycharmProjects\pythonProject\otherTF"
WARM_MAPS = r"C:\Users\777\PycharmProjects\pythonProject\WarmMaps"
os.makedirs(WARM_MAPS, exist_ok=True)

print("🔍 Сканируем папки с исходными данными...")
h1_files = glob.glob(os.path.join(OTHER_TF, "*_H1.txt"))
d1_files = glob.glob(os.path.join(OTHER_TF, "*_D1.txt"))
tickers = set(os.path.basename(f).split("_")[0] for f in h1_files + d1_files)
print(f"Найдено тикеров: {len(tickers)} -> {sorted(tickers)}")

if not tickers:
    print("❗ Нет файлов для обработки. Выход.")
    exit()

# 1) Формируем файл RESULT_HEAT_MAP.xlsx с отдельными листами D1 и H1 у каждого тикера
combined_path = os.path.join(WARM_MAPS, "RESULT_HEAT_MAP.xlsx")
print(f"📂 Готовим файл: {combined_path}")

pivots = {}

with pd.ExcelWriter(combined_path, engine="openpyxl") as writer:
    for ticker in sorted(tickers):
        print(f"\n🔵 Обрабатываем тикер: {ticker}")
        h1_file = next((f for f in h1_files if os.path.basename(f).startswith(ticker)), None)
        d1_file = next((f for f in d1_files if os.path.basename(f).startswith(ticker)), None)
        if not h1_file or not d1_file:
            print(f"  ❌ Пропуск: отсутствуют H1 или D1 файлы для {ticker}")
            continue

        # Читаем D1
        print(f"  📥 Читаем D1: {os.path.basename(d1_file)}")
        d1 = pd.read_csv(
            d1_file,
            skiprows=1,
            names=['ticker','per','date','time','open','high','low','close','vol']
        )
        for col in ['open','high','low','close','vol']:
            d1[col] = pd.to_numeric(d1[col], errors='coerce')
        d1.dropna(subset=['open','high','low','close'], inplace=True)
        d1['datetime'] = pd.to_datetime(
            d1['date'].astype(str).str.zfill(8) + d1['time'].astype(str).str.zfill(6),
            format='%Y%m%d%H%M%S'
        )
        d1.sort_values('datetime', inplace=True)
        d1['ampl_pct'] = 2 * (d1['high'] - d1['low']) / (d1['high'] + d1['low']) * 100
        print(f"    ✔ D1 обработан: {len(d1)} строк")

        # Читаем H1
        print(f"  📥 Читаем H1: {os.path.basename(h1_file)}")
        h1 = pd.read_csv(
            h1_file,
            skiprows=1,
            names=['ticker','per','date','time','open','high','low','close','vol']
        )
        for col in ['open','high','low','close','vol']:
            h1[col] = pd.to_numeric(h1[col], errors='coerce')
        h1.dropna(subset=['open','high','low','close'], inplace=True)
        h1['datetime'] = pd.to_datetime(
            h1['date'].astype(str).str.zfill(8) + h1['time'].astype(str).str.zfill(6),
            format='%Y%m%d%H%M%S'
        )
        h1.sort_values('datetime', inplace=True)
        h1['ampl_pct'] = 2 * (h1['high'] - h1['low']) / (h1['high'] + h1['low']) * 100
        h1['hour'] = h1['datetime'].dt.strftime('%H:%M')
        h1['weekday'] = (h1['datetime'] - pd.Timedelta(hours=3)).dt.weekday
        h1['weekday_name'] = h1['weekday'].map({0:'Пн', 1:'Вт', 2:'Ср', 3:'Чт', 4:'Пт', 5:'Сб', 6:'Вс'})
        print(f"    ✔ H1 обработан: {len(h1)} строк")

        # Формируем pivot-таблицу
        hour_list = [f"{h:02d}:00" for h in list(range(3,24)) + list(range(0,3))]
        print("  🔄 Формируем pivot-таблицу")
        pivot = h1.pivot_table(
            index='weekday_name',
            columns='hour',
            values='ampl_pct',
            aggfunc='mean'
        ).reindex(index=['Пн','Вт','Ср','Чт','Пт','Сб','Вс'], columns=hour_list)
        print("    ✔ Pivot создан")

        # Добавляем строки и столбцы со Средним и Медианой
        print("  ➕ Добавляем строки/столбцы Среднее/Медиана")
        data_cols = len(hour_list)

        col_means = pivot.iloc[:, :data_cols].mean()
        col_meds  = pivot.iloc[:, :data_cols].median()

        pivot.loc['Среднее'] = col_means
        pivot.loc['Медиана'] = col_meds

        pivot['Среднее'] = pivot.iloc[:, :data_cols].mean(axis=1)
        pivot['Медиана'] = pivot.iloc[:, :data_cols].median(axis=1)

        overall_mean   = pivot.iloc[:-2, :data_cols].values.mean()
        overall_median = np.median(pivot.iloc[:-2, :data_cols].values.flatten())
        pivot.at['Среднее', 'Среднее'] = overall_mean
        pivot.at['Медиана', 'Медиана'] = overall_median
        print("    ✔ Среднее/Медиана добавлены")

        pivots[ticker] = pivot.copy()

        # Записываем листы D1 и H1
        d1_sheet = f"{ticker}_D1"
        h1_sheet = f"{ticker}_H1"
        print(f"  💾 Записываем листы: {d1_sheet}, {h1_sheet}")
        d1[['datetime','open','high','low','close','ampl_pct']].to_excel(writer, sheet_name=d1_sheet, index=False)
        pivot.to_excel(writer, sheet_name=h1_sheet)
        print("    ✔ Листы записаны")

print("\n✅ RESULT_HEAT_MAP.xlsx создан")

# 2) Применяем условное форматирование (светло-синий→белый→красный) в RESULT_HEAT_MAP.xlsx
print("🔧 Наложение условного форматирования в RESULT_HEAT_MAP.xlsx...")
wb = load_workbook(combined_path)
for ticker, pivot in pivots.items():
    h1_sheet = f"{ticker}_H1"
    if h1_sheet not in wb.sheetnames:
        print(f"  ⚠️ Лист {h1_sheet} не найден, пропуск")
        continue
    ws = wb[h1_sheet]
    data_rows = 7
    data_cols = len(hour_list)
    start_col = 2
    start_row = 2

    # 1) Основная область
    data_start = f"{col_letter(start_col)}{start_row}"
    data_end   = f"{col_letter(start_col + data_cols - 1)}{start_row + data_rows - 1}"
    data_range = f"{data_start}:{data_end}"
    rule_main = ColorScaleRule(
        start_type='min', start_color='ADD8E6',   # светло-синий = минимум
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'        # красный = максимум
    )
    ws.conditional_formatting.add(data_range, rule_main)

    # 2) Строки «Среднее» и «Медиана»
    row_mean = start_row + data_rows
    row_med  = row_mean + 1
    summary_row1 = f"{col_letter(start_col)}{row_mean}:{col_letter(start_col + data_cols + 1)}{row_mean}"
    summary_row2 = f"{col_letter(start_col)}{row_med}:{col_letter(start_col + data_cols + 1)}{row_med}"
    rule_rows = ColorScaleRule(
        start_type='min', start_color='ADD8E6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws.conditional_formatting.add(summary_row1, rule_rows)
    ws.conditional_formatting.add(summary_row2, rule_rows)

    # 3) Столбцы «Среднее» и «Медиана»
    col_mean = start_col + data_cols
    col_med  = col_mean + 1
    letter_mean = col_letter(col_mean)
    letter_med  = col_letter(col_med)
    summary_col1 = f"{letter_mean}{start_row}:{letter_mean}{start_row + data_rows - 1}"
    summary_col2 = f"{letter_med}{start_row}:{letter_med}{start_row + data_rows - 1}"
    rule_cols = ColorScaleRule(
        start_type='min', start_color='ADD8E6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws.conditional_formatting.add(summary_col1, rule_cols)
    ws.conditional_formatting.add(summary_col2, rule_cols)

wb.save(combined_path)
print("✅ Условное форматирование для RESULT_HEAT_MAP.xlsx завершено")

# 3) Собираем все pivot-таблицы подряд в один лист SUMMARY_TICKERS.xlsx
summary_all_path = os.path.join(WARM_MAPS, "SUMMARY_TICKERS.xlsx")
print(f"\n📂 Собираем все pivot-таблицы в один лист: {summary_all_path}")
with pd.ExcelWriter(summary_all_path, engine="openpyxl") as writer:
    sheet_name = "All_H1"
    current_row = 0
    for ticker, pivot in pivots.items():
        # Заголовок с именем тикера
        df_header = pd.DataFrame([[ticker]], columns=[f"{ticker}_H1"])
        df_header.to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False)
        current_row += 1

        # Записываем pivot-таблицу под этим заголовком
        pivot.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0)
        rows, cols = pivot.shape
        current_row += rows + 2  # +1 строка pivot и +1 пустая строка

print("✅ SUMMARY_TICKERS.xlsx создан (без форматирования)")

# 4) Применяем условное форматирование (светло-синий→белый→красный) в SUMMARY_TICKERS.xlsx
print("🔧 Применяем условное форматирование в SUMMARY_TICKERS.xlsx...")
wb2 = load_workbook(summary_all_path)
ws2 = wb2["All_H1"]

current_row = 1
for ticker, pivot in pivots.items():
    rows, total_cols = pivot.shape
    data_cols = total_cols - 2
    start_col = 2
    start_row = current_row + 1

    # 4.1) Основной блок
    data_start = f"{col_letter(start_col)}{start_row}"
    data_end   = f"{col_letter(start_col + data_cols - 1)}{start_row + 6}"
    data_range = f"{data_start}:{data_end}"
    rule_main2 = ColorScaleRule(
        start_type='min', start_color='ADD8E6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws2.conditional_formatting.add(data_range, rule_main2)

    # 4.2) Строки «Среднее» и «Медиана»
    row_mean = start_row + 7
    row_med  = row_mean + 1
    summary_row1 = f"{col_letter(start_col)}{row_mean}:{col_letter(start_col + data_cols + 1)}{row_mean}"
    summary_row2 = f"{col_letter(start_col)}{row_med}:{col_letter(start_col + data_cols + 1)}{row_med}"
    rule_rows2 = ColorScaleRule(
        start_type='min', start_color='ADD8E6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws2.conditional_formatting.add(summary_row1, rule_rows2)
    ws2.conditional_formatting.add(summary_row2, rule_rows2)

    # 4.3) Столбцы «Среднее» и «Медиана»
    col_mean = start_col + data_cols
    col_med  = col_mean + 1
    letter_mean = col_letter(col_mean)
    letter_med  = col_letter(col_med)
    summary_col1 = f"{letter_mean}{start_row}:{letter_mean}{start_row + 6}"
    summary_col2 = f"{letter_med}{start_row}:{letter_med}{start_row + 6}"
    rule_cols2 = ColorScaleRule(
        start_type='min', start_color='ADD8E6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws2.conditional_formatting.add(summary_col1, rule_cols2)
    ws2.conditional_formatting.add(summary_col2, rule_cols2)

    current_row += rows + 2 + 1

wb2.save(summary_all_path)
print("✅ Условное форматирование для SUMMARY_TICKERS.xlsx завершено")

# 5) Наносим толстые рамки для строк «Среднее» и «Медиана», обводим и столбцы, и строки;
#    а также заливаем желтым цветом самую правую нижнюю ячейку каждого листа

# Подготовка «толстой» рамки и жёлтой заливки
thick = Side(border_style="thick", color="000000")
border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)
yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

# Работаем над SUMMARY_TICKERS.xlsx
print("🔧 Наносим рамки и заливаем жёлтым в SUMMARY_TICKERS.xlsx...")
wb3 = load_workbook(summary_all_path)

for sheetname in wb3.sheetnames:
    ws = wb3[sheetname]
    print(f"Обрабатываем лист: {sheetname}")
    max_row = ws.max_row
    max_col = ws.max_column

    # 5.1) Обводим строки, где есть «Среднее» или «Медиана»
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        found = False
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() in ("Среднее", "Медиана"):
                found = True
                break
        if not found:
            continue

        row_idx = row[0].row
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).border = border_thick

    # 5.2) Обводим столбцы, где заголовок — «Среднее» или «Медиана»
    #      Предполагаем, что заголовок стоит в первой строке (или первой непустой).
    #      Если формат иной, подправьте min_row.
    for col_idx in range(1, max_col + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if isinstance(header_val, str) and header_val.strip() in ("Среднее", "Медиана"):
            for row_idx in range(1, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).border = border_thick

wb3.save(summary_all_path)
print("✅ Рамки и заливка в SUMMARY_TICKERS.xlsx завершены")

